import openpyxl
from openpyxl.utils import get_column_letter
import config


def read_tracker_schema(excel_path: str) -> dict:
    """
    Read the Excel tracker and return its schema.

    Returns:
        {
            "header_row": int,
            "data_start_row": int,
            "data_end_row": int,
            "columns": {col_index: header_text, ...},
            "section_headers": {col_index: section_name, ...},  # from row 2
            "merged_cells": list of merged ranges,
        }
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Find the header row: the row with the most non-empty cells in the search range
    header_row = _find_header_row(ws)

    # Read column headers from the header row
    columns = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            columns[col] = str(val).strip()

    # Read section headers (typically row 2 in these trackers)
    section_headers = {}
    if header_row > 1:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if val:
                section_headers[col] = str(val).strip()

    # Find data range
    data_start_row = header_row + 1
    data_end_row = ws.max_row

    # Get merged cells
    merged = [str(m) for m in ws.merged_cells.ranges]

    wb.close()
    return {
        "header_row": header_row,
        "data_start_row": data_start_row,
        "data_end_row": data_end_row,
        "columns": columns,
        "section_headers": section_headers,
        "merged_cells": merged,
    }


def read_existing_data(excel_path: str, schema: dict = None) -> list[dict]:
    """
    Read existing data rows from the tracker.

    Returns a list of dicts, one per row, with column index as keys.
    Only includes cells that have values.
    """
    if schema is None:
        schema = read_tracker_schema(excel_path)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    rows = []
    for row_num in range(schema["data_start_row"], schema["data_end_row"] + 1):
        row_data = {"_row_num": row_num}
        has_data = False
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row_num, column=col).value
            if val is not None:
                row_data[col] = val
                has_data = True
        if has_data:
            rows.append(row_data)

    wb.close()
    return rows


def get_column_headers_list(schema: dict) -> list[str]:
    """Return a formatted list of column headers for LLM prompts."""
    headers = []
    for col_idx in sorted(schema["columns"].keys()):
        header = schema["columns"][col_idx]
        letter = get_column_letter(col_idx)
        headers.append(f"Column {letter} ({col_idx}): {header}")
    return headers


def get_empty_columns_for_row(existing_data: list[dict], row_index: int, schema: dict) -> list[int]:
    """Return column indices that are empty for a given data row."""
    if row_index >= len(existing_data):
        return list(schema["columns"].keys())

    row = existing_data[row_index]
    empty = []
    for col_idx in schema["columns"]:
        if col_idx not in row or row[col_idx] is None:
            empty.append(col_idx)
    return empty


def _find_header_row(ws) -> int:
    """Find the row with the most non-empty cells (the actual header row)."""
    best_row = 1
    best_count = 0

    search_range = min(config.HEADER_ROW_SEARCH_RANGE, ws.max_row)
    for row in range(1, search_range + 1):
        count = sum(
            1 for col in range(1, ws.max_column + 1)
            if ws.cell(row=row, column=col).value is not None
        )
        if count > best_count:
            best_count = count
            best_row = row

    return best_row
