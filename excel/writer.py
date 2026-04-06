from __future__ import annotations

import re
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def _parse_column_key(key: str) -> int | None:
    """
    Parse a column identifier like 'Column B (2)' into a column index.
    Also handles plain column letters like 'B' or plain numbers like '2'.
    """
    # Match "Column X (N)" pattern
    match = re.match(r"Column\s+([A-Z]+)\s*\((\d+)\)", key)
    if match:
        return int(match.group(2))

    # Match plain column letter
    match = re.match(r"^([A-Z]{1,3})$", key.strip())
    if match:
        try:
            return column_index_from_string(match.group(1))
        except ValueError:
            return None

    # Match plain number
    match = re.match(r"^(\d+)$", key.strip())
    if match:
        return int(match.group(1))

    return None


def _coerce_value(value, existing_cell_value=None):
    """Convert extracted string values to appropriate Python types."""
    if value is None or value == "":
        return None

    s = str(value).strip()
    if s.upper() in ("N/A", "NA", "NONE", "NULL"):
        return "N/A"

    # Preserve currency-prefixed strings (e.g. "INR 1,27,97,788.63", "USD 500.00")
    if re.match(r"^[A-Z]{2,3}\s", s):
        return s

    # Try numeric
    try:
        # Remove commas
        cleaned = s.replace(",", "")
        if "." in cleaned:
            return float(cleaned)
        return int(cleaned)
    except ValueError:
        pass

    return s


def write_to_tracker(
    template_path: str,
    output_path: str,
    merged_data: dict,
    schema: dict,
    existing_data: list[dict],
    overwrite_existing: bool = False,
) -> str:
    """
    Write extracted data to the Excel tracker.

    Args:
        template_path: Path to the original Excel file
        output_path: Path to save the filled Excel file
        merged_data: {"shared_row": {...}, "per_vehicle_rows": [...]}
        schema: From excel.reader.read_tracker_schema()
        existing_data: From excel.reader.read_existing_data()
        overwrite_existing: If True, overwrite pre-filled cells

    Returns the output path.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    shared = merged_data.get("shared_row", {})
    per_vehicle = _dedup_per_vehicle_list(merged_data.get("per_vehicle_rows", []))

    data_start = schema["data_start_row"]

    if per_vehicle:
        # Write one row per vehicle: shared columns + per-vehicle columns merged
        for i, vehicle_data in enumerate(per_vehicle):
            row_num = data_start + i
            # Shared data first, then vehicle-specific data overwrites
            _write_row_data(ws, row_num, shared, overwrite_existing)
            _write_row_data(ws, row_num, vehicle_data, overwrite_existing)
    else:
        # No per-vehicle data — write shared data to a single row
        _write_row_data(ws, data_start, shared, overwrite_existing)

    wb.save(output_path)
    wb.close()
    return output_path


def _dedup_per_vehicle_list(per_vehicle: list[dict]) -> list[dict]:
    """Final safety-net merge by BoE/vehicle before writing to Excel."""

    def _is_blank(value) -> bool:
        return str(value).strip().upper() in ("", "NONE", "N/A", "NULL")

    def _find_boe(row: dict) -> str | None:
        for key, value in row.items():
            key_lower = str(key).lower()
            if "boe" in key_lower or "bill of entry" in key_lower or "be number" in key_lower:
                if not _is_blank(value):
                    return str(value).strip().upper()
        return None

    def _find_vehicle(row: dict) -> str | None:
        for key, value in row.items():
            if "vehicle" in str(key).lower() and not _is_blank(value):
                return str(value).strip().upper()
        return None

    def _find_invoice(row: dict) -> str | None:
        for key, value in row.items():
            if "invoice" in str(key).lower() and not _is_blank(value):
                return str(value).strip().upper()
        return None

    def _merge_rows(base: dict, incoming: dict) -> dict:
        merged_row = dict(base)
        for key, value in incoming.items():
            if key.startswith("_") or _is_blank(value):
                continue
            if key not in merged_row or _is_blank(merged_row[key]):
                merged_row[key] = value
        return merged_row

    keyed: dict[str, dict] = {}
    seen_hashes: set[str] = set()
    unique: list[dict] = []

    for row in per_vehicle:
        boe = _find_boe(row)
        vehicle = _find_vehicle(row)
        invoice = _find_invoice(row)

        key = None
        if boe:
            key = f"boe::{boe}"
        elif vehicle and invoice:
            key = f"vehicle_invoice::{vehicle}||{invoice}"
        elif vehicle:
            key = f"vehicle::{vehicle}"

        if key:
            if key in keyed:
                keyed[key] = _merge_rows(keyed[key], row)
            else:
                keyed[key] = dict(row)
            continue

        h = "|".join(
            f"{k}={str(row[k]).strip().upper()}" for k in sorted(row.keys())
            if not _is_blank(row[k])
        )
        if h in seen_hashes:
            continue
        seen_hashes.add(h)
        unique.append(dict(row))

    unique.extend(keyed.values())
    return unique


def _write_row_data(ws, row_num: int, data: dict, overwrite: bool):
    """Write a dict of column values to a specific row."""
    for key, value in data.items():
        if key.startswith("_"):
            continue
        col_idx = _parse_column_key(key)
        if col_idx is None:
            continue

        cell = ws.cell(row=row_num, column=col_idx)
        if cell.value is not None and not overwrite:
            continue  # Skip pre-filled cells

        coerced = _coerce_value(value, cell.value)
        if coerced is not None:
            cell.value = coerced


def _write_per_vehicle_data(
    ws,
    per_vehicle: list[dict],
    existing_data: list[dict],
    data_start: int,
    schema: dict,
    overwrite: bool,
):
    """Match per-vehicle data to existing rows using vehicle number."""
    # Build a lookup of vehicle_no -> row_num from existing data
    vehicle_to_row = {}
    for row_data in existing_data:
        row_num = row_data.get("_row_num")
        # Vehicle number is typically in column 63 (BK)
        vehicle_no = row_data.get(63)
        if vehicle_no and row_num:
            vehicle_to_row[str(vehicle_no).strip().upper()] = row_num

    for vehicle_data in per_vehicle:
        # Find the vehicle number in the extracted data
        vehicle_no = None
        for key, value in vehicle_data.items():
            if "vehicle" in key.lower() or key == "vehicle_no":
                vehicle_no = str(value).strip().upper()
                break

        if vehicle_no and vehicle_no in vehicle_to_row:
            row_num = vehicle_to_row[vehicle_no]
            _write_row_data(ws, row_num, vehicle_data, overwrite)
        else:
            # If no match, write to the next available row
            # Find the first row without this data
            pass  # Skip unmatched vehicles - user can review
